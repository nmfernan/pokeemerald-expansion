#Tool for creating new gen.h files for emerald expansion
#Takes information from excel sheet and prepares .h file for custom pokemon
import os
import glob
import re
import shutil
import openpyxl as pyxl

from openpyxl.workbook import Workbook
from openpyxl import load_workbook

Debug = 0
Clean = 1

FrontSprite = "front.png"
BackSprite = "back.png"
Icon = "icon.png"
InputFolderPath = "Sprites-MASTER/"
FrontSpritePath = "FrontSprite"
FrontSpriteShinyPath = "FrontSpriteShiny"
BackSpritePath = "BackSprite"
BackSpriteShinyPath = "BackSpriteShiny"
IconPath = "Icon"
OutputFolderPath = "graphics"

SpecialCases = ["NIDORAN_M",#Nidoran_m
                "NIDORAN_F",#Nidoran_f
                "MR_MIME",#mr mime
                "MR_JEST",#mr jest
                "MR_FOOL"]#mr fool

#Globals for making header, opening data, debug prints, etc
PkmnData = load_workbook('pkmndata.xlsx')
PkmnDataFile = PkmnData['sanity-data']

# print(f"FrontSprite Dir: {os.path.isdir(InputFolderPath + FrontSpritePath)}")
# print(f"BackSprite Dir: {os.path.isdir(InputFolderPath + BackSpritePath)}")
# print(f"FrontSpriteShiny Dir: {os.path.isdir(InputFolderPath + FrontSpriteShinyPath)}")
# print(f"BackSpriteShiny Dir: {os.path.isdir(InputFolderPath + BackSpriteShinyPath)}")
# print(f"Icon Dir: {os.path.isdir(InputFolderPath + IconPath)}\n")
# os.makedirs(OutputFolderPath, exist_ok = True)

print(f"First row for species {PkmnDataFile.min_row}")
print(f"Last row for species {PkmnDataFile.max_row}")
print(f"First column of species-data {PkmnDataFile.min_column}")
print(f"Last column of species-data  {PkmnDataFile.max_column}")    

if Clean == 1:
    for row in PkmnDataFile.iter_rows(min_row=2, max_row=PkmnDataFile.max_row, min_col=1, max_col=1):
        print(row[0].value)
        for path in os.listdir():
            files = glob.glob(f"{path}/*smol")
            #print(path)
            if path == row[0].value.lower():
                for file in files:
                    try:
                        os.remove(file)
                        print("deleted")
                    except:
                        print("error deleting")

    for row in PkmnDataFile.iter_rows(min_row=2, max_row=PkmnDataFile.max_row, min_col=1, max_col=1):
        print(row[0].value)
        for path in os.listdir():
            files = glob.glob(f"{path}/*pal")
            #print(path)
            if path == row[0].value.lower():
                for file in files:
                    try:
                        os.remove(file)
                        print("deleted")
                    except:
                        print("error deleting")
    
    for row in PkmnDataFile.iter_rows(min_row=2, max_row=PkmnDataFile.max_row, min_col=1, max_col=1):
        print(row[0].value)
        for path in os.listdir():
            files = glob.glob(f"{path}/*4bpp")
            #print(path)
            if path == row[0].value.lower():
                for file in files:
                    try:
                        os.remove(file)
                        print("deleted")
                    except:
                        print("error deleting")

#         for path in os.listdir():
#             if path == row[0].value.lower():
#                 print(row[0].value)
#                 for file in os.listdir(f"./{path}"):
#                     print(file)                    
#                     if "pal" in file or "4bpp" in file:
#                         try:
#                             os.remove("TEST.TXT")
#                             print("deleted")
#                         except:
# #                            print("error deleting")
#                             break
#                         print(file)
# 
# 