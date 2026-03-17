#Tool for creating new gen.h files for emerald expansion
#Takes information from excel sheet and prepares .h file for custom pokemon
import time
import openpyxl as pyxl

from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

start = time.time()

#Globals for execution
Debug = 1
WriteOrAdd = 'w'
GenName = "Evo"
FileName = "evo_families.h"
PkmnData = load_workbook('pkmndata.xlsx')
PkmnDataFile = PkmnData['sanity-data']
OnlyNewSpecies = 0

CurrentFamily = ""
CurrentSpecies = ""
CurrentSpeciesCptl = ""
Attribute = ""

#Header Description for files, etc
Header = """//gen file for """ + GenName + """
#ifdef __INTELLISENSE__
const struct SpeciesInfo gSpeciesInfo""" + GenName + """[] =
{
#endif
"""
#Header Description for files, etc
Footer ="""#ifdef __INTELLISENSE__
};
#endif
"""

#Definiton of species information in array format for adding to .h file when iterating through data values
#Just a sanity check to print at the start
if Debug:
    SpeciesStructAttributes = []
    for row in PkmnDataFile.iter_rows(min_row=1, max_row=1, min_col=1, max_col=PkmnDataFile.max_column):
        for data in row:
            SpeciesStructAttributes.append(str(data.value))
            print(data.value)

if OnlyNewSpecies != 1:
    with open(FileName, WriteOrAdd) as file:
        #Print high level information about datafile being accessed
        if Debug:
            print(f"First row for species {PkmnDataFile.min_row}")
            print(f"Last row for species {PkmnDataFile.max_row}")
            print(f"First column of species-data {PkmnDataFile.min_column}")
            print(f"Last column of species-data  {PkmnDataFile.max_column}")
            print(f"First column of tutor-data #{PkmnDataFile.min_column}, Letter:{get_column_letter(PkmnDataFile.min_column)}")
            print(f"Last column of tutor-data  #{PkmnDataFile.max_column}, Letter:{get_column_letter(PkmnDataFile.max_column)}")    
        
        #Write top level information to species file
        file.write(Header)           

        #Start from second row so you do not grab data headers
        #for species in PkmnDataFile.iter_rows(min_row=2, max_row=310, min_col=PkmnDataFile.min_column, max_col=PkmnDataFile.max_column):
        for species in PkmnDataFile.iter_rows(min_row=2, max_row=PkmnDataFile.max_row, min_col=PkmnDataFile.min_column, max_col=PkmnDataFile.max_column):
            CurrentSpecies = species[0].value
            CurrentSpeciesCptl = CurrentSpecies.capitalize()
            
            if species[PkmnDataFile.max_column-1].value == 1:#species tuple is 0 indexed; maxcol is 1 indexed
                CurrentFamily = species[0].value
                
                if CurrentFamily == "NIDORAN_F":
                    strip = CurrentFamily.rfind("_")
                    file.write(f"\n#if P_FAMILY_{CurrentFamily[:strip]}\n")
                else:
                    file.write(f"\n#if P_FAMILY_{CurrentFamily}\n")
                
                if Debug: print(f"New Species Found!: {CurrentFamily}")    
            file.write(f"\t[SPECIES_{CurrentSpecies}] =\n")
            file.write("\t{\n")
            
            #step through each element of the species
            for data in species:
                Attribute = PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value
                
                if Attribute == ".types":
                    types = data.value.split(',')
                    type1 = types[0]
                    type2 = types[1]
                    if type1 == type2: #Check for mono typing
                        file.write(f"\t\t.types = MON_TYPES(TYPE_{type1}),\n")
                    else:
                        file.write(f"\t\t.types = MON_TYPES(TYPE_{type1}, TYPE_{type2}),\n")    
                
                elif Attribute == ".eggGroups":
                    types = data.value.split(',')
                    type1 = types[0]
                    type2 = types[1]
                    if type1 == type2: #Check for single egg group
                        file.write(f"\t\t.eggGroups = MON_EGG_GROUPS(EGG_GROUP_{type1}),\n")
                    else:
                        file.write(f"\t\t.eggGroups = MON_EGG_GROUPS(EGG_GROUP_{type1}, EGG_GROUP_{type2}),\n")    
                
                elif Attribute == ".abilities":
                    ability = data.value.split(',')
                    ability1 = f"ABILITY_{ability[0]}"
                    ability2 = f"ABILITY_{ability[1]}"
                    ability3 = f"ABILITY_{ability[2]}"
                    file.write(f"\t\t.abilities = {{ {ability1}, {ability2}, {ability3} }},\n")
                
                elif Attribute == ".bodyColor":
                    file.write(f"\t\t.bodyColor = BODY_COLOR_{data.value},\n")
                
                elif Attribute == ".natDexNum":
                    file.write(f"\t\t.natDexNumber = NATIONAL_DEX_{CurrentSpecies}\n")
                
                elif Attribute == ".speciesName":
                    file.write(f"\t\t{Attribute} = _(\"{CurrentSpeciesCptl}\"),\n")
                
                elif Attribute == ".categoryName":
                    Category = data.value.capitalize()
                    file.write(f"\t\t{Attribute} = _(\"{Category}\"),\n")
                
                elif Attribute == ".description":
                    file.write(f"\t\t{Attribute} = COMPOUD_STRING(\"{data.value}\"),\n")
                
                elif Attribute == ".frontPic":
                    file.write(f"\t\t{Attribute} = gMonFrontPic_{CurrentSpeciesCptl},\n")
                
                elif Attribute == ".frontPicSize":
                    file.write(f"\t\t{Attribute} = MON_COORDS_SIZE({data.value}),\n")
                
                elif Attribute == ".backPic":
                    file.write(f"\t\t{Attribute} = gMonBackPic_{CurrentSpeciesCptl},\n")
                
                elif Attribute == ".backPicSize":
                    file.write(f"\t\t{Attribute} = MON_COORDS_SIZE({data.value}),\n")
                
                elif Attribute == ".palette":
                    file.write(f"\t\t{Attribute} = gMonPalette_{CurrentSpeciesCptl},\n")
                
                elif Attribute == ".shinyPalette":
                    file.write(f"\t\t{Attribute} = gMonShinyPalette_{CurrentSpeciesCptl},\n")
                
                elif Attribute == "FOOTPRINT":
                    file.write(f"\t\t{Attribute}({CurrentSpeciesCptl})\n")    

                elif Attribute == ".iconSprite":
                    file.write(f"\t\t{Attribute} = gMonIcon_{CurrentSpeciesCptl},\n")
                
                elif Attribute == ".levelUpLearnset":
                    #file.write(f"\t\t{Attribute} = s{CurrentSpeciesCptl}LevelUpLearnset,\n")
                    continue
                elif Attribute == ".teachableLearnset":
                    #file.write("\t\t{Attribute} = s" + CurrentSpeciesCptl  + "TeachableLearnset,\n")
                    continue
    #             elif PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".eggMoveLearnset":
    #                 CurrentSpeciesCptl = PkmnDataFile.cell(row = data.row, column = PkmnDataFile.min_column).value
    #                 CurrentSpeciesCptl = CurrentSpeciesCptl[0] + CurrentSpeciesCptl[1:len(CurrentSpeciesCptl)].lower()
    #                 file.write("\t\t.eggMoveLearnset = s" + CurrentSpeciesCptl  + "TeachableLearnset,\n")
#                elif Attribute == ".evolutions" and data.value != None:
#                    file.write(f"\t\t{Attribute} = EVOLUTION({{EVO_LEVEL, {data.value}, SPECIES_{PkmnDataFile.cell(data.row + 1, PkmnDataFile.min_column).value}}}),\n")
                elif Attribute == ".evolutions" and data.value != None:
                    if data.value.find("STONE") > 0:
                        file.write(f"\t\t{Attribute} = EVOLUTION({{EVO_ITEM, {data.value}, SPECIES_{PkmnDataFile.cell(data.row + 1, PkmnDataFile.min_column).value}}}),\n")
                    elif CurrentSpecies == "BEEBRUTE": #brute force the beebrute interaction for now
                        file.write(f"\t\t{Attribute} = EVOLUTION({{EVO_LEVEL, {data.value}, SPECIES_SEPISTRIKE}},\n\t\t\t\t{{EVO_LEVEL, {data.value}, SPECIES_DRONARCH, CONDITIONS({{IF_GENDER, MON_MALE}})}}),\n")
                    elif CurrentSpecies == "EEVEE":
                        file.write(f"\t\t{Attribute} = EVOLUTION({{EVO_ITEM, ITEM_THUNDER_STONE, SPECIES_JOLTEON}},{{EVO_ITEM, ITEM_WATER_STONE, SPECIES_VAPOREON}},{{EVO_ITEM, ITEM_FIRE_STONE, SPECIES_FLAREON}}),\n")
                    else:
                        file.write(f"\t\t{Attribute} = EVOLUTION({{EVO_LEVEL, {data.value}, SPECIES_{PkmnDataFile.cell(data.row + 1, PkmnDataFile.min_column).value}}}),\n")
                              
                elif Attribute == "newspecies" or Attribute == ".natDexNeeded":
                    continue
                
                elif not data.value:
                    continue
                
                else:
                    file.write(f"\t\t{Attribute} = {data.value},\n")
                
            #close this species
            file.write("\t\t},\n\n")
            
            if PkmnDataFile.cell(species[0].row + 1, PkmnDataFile.max_column).value == 1:#check if new species is starting in next row
                file.write(f"#endif//P_FAMILY_{CurrentFamily}\n")
        
        file.write(f"#endif//P_FAMILY_{CurrentFamily}\n\n")
    
        end = time.time()
        print(f"Time: {end - start} seconds")
        
        file.write(Footer)
        file.write("//end of program")


if OnlyNewSpecies:
    with open(FileName, WriteOrAdd) as file:
        #Print high level information about datafile being accessed
        if Debug:
            print(f"First row for species {PkmnDataFile.min_row}")
            print(f"Last row for species {PkmnDataFile.max_row}")
            print(f"First column of species-data {PkmnDataFile.min_column}")
            print(f"Last column of species-data  {PkmnDataFile.max_column}")
            print(f"First column of tutor-data #{PkmnDataFile.min_column}, Letter:{get_column_letter(PkmnDataFile.min_column)}")
            print(f"Last column of tutor-data  #{PkmnDataFile.max_column}, Letter:{get_column_letter(PkmnDataFile.max_column)}")    
        
        #Write top level information to species file
        file.write(Header +  "\n")           

        #Begin writing species information to .h file
        #Start from second row so you do not grab data headers
        for species in PkmnDataFile.iter_rows(min_row=2, max_row=6, min_col=PkmnDataFile.min_column, max_col=PkmnDataFile.max_column):
        #for species in PkmnDataFile.iter_rows(min_row=2, max_row=PkmnDataFile.max_row, min_col=PkmnDataFile.min_column, max_col=PkmnDataFile.max_column):
            
            #step through each element of the species
            if species[PkmnDataFile.max_column-2].value == 1:#species tuple is 0 indexed; maxcol is 1 indexed
                print("New Species Found!: " + species[PkmnDataFile.min_column-1].value)
                file.write("\t[SPECIES_" + str(species[PkmnDataFile.min_column-1].value) + "] =\n")
                file.write("\t{\n")
            
                for data in species:
                    if PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".types":
                        types = data.value.split(',')
                        type1 = types[0]
                        type2 = types[1]
                        if type1 == type2: #Check for single typing
                            file.write("\t\t.types = MON_TYPES(TYPE_" + type1 + "),\n")
                        else:
                            file.write("\t\t.types = MON_TYPES(TYPE_" + type1 + ", TYPE_"+ type2 + "),\n")    
                    elif PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".eggGroups":
                        types = data.value.split(',')
                        type1 = types[0]
                        type2 = types[1]
                        if type1 == type2: #Check for single egg group
                            file.write("\t\t.eggGroups = MON_EGG_GROUPS(EGG_GROUP_" + type1 + "),\n")
                        else:
                            file.write("\t\t.eggGroups = MON_EGG_GROUPS(EGG_GROUP_" + type1 + ", EGG_GROUP_"+ type2 + "),\n")    
                    elif PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".abilities":
                        types = data.value.split(',')
                        type1 = "ABILITY_" + types[0]
                        type2 = "ABILITY_" + types[1]
                        type3 = "ABILITY_" + types[2]
                        file.write("\t\t.abilities = { " + type1 + ", " + type2 + " , " + type3 + " },\n")
                    elif PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".bodyColor":
                        file.write("\t\t.bodyColor = BODY_COLOR_" + data.value + ",\n")
                    elif PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".speciesName":
                        CurrentSpeciesCptl = data.value
                        CurrentSpeciesCptl = CurrentSpeciesCptl[0] + CurrentSpeciesCptl[1:len(CurrentSpeciesCptl)].lower()
                        file.write("\t\t" + SpeciesStructAttributes[data.column-1] + " = _(\"" + CurrentSpeciesCptl + "\"),\n")
                    elif PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".categoryName":
                        CurrentSpeciesCptl = data.value
                        CurrentSpeciesCptl = CurrentSpeciesCptl[0] + CurrentSpeciesCptl[1:len(CurrentSpeciesCptl)].lower()
                        file.write("\t\t" + SpeciesStructAttributes[data.column-1] + " = _(\"" + CurrentSpeciesCptl + "\"),\n")
                    elif PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".description":
        #                file.write("\t\t.description = COMPOUD_STRING(\n\t\t\t\"" + data.value + "\"),\n")
                        continue
                    elif PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".frontPic":
                        CurrentSpeciesCptl = PkmnDataFile.cell(row = data.row, column = PkmnDataFile.min_column).value
                        CurrentSpeciesCptl = CurrentSpeciesCptl[0] + CurrentSpeciesCptl[1:len(CurrentSpeciesCptl)].lower()
                        file.write("\t\t.frontPic = gMonFrontPic_" + CurrentSpeciesCptl  + ",\n")
                    elif PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".frontPicSize":
                        file.write("\t\t.frontPicSize = MON_COORDS_SIZE(" + data.value + "),\n")
        #                continue
                    elif PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".backPic":
                        CurrentSpeciesCptl = PkmnDataFile.cell(row = data.row, column = PkmnDataFile.min_column).value
                        CurrentSpeciesCptl = CurrentSpeciesCptl[0] + CurrentSpeciesCptl[1:len(CurrentSpeciesCptl)].lower()
                        file.write("\t\t.backPic = gMonBackPic_" + CurrentSpeciesCptl  + ",\n")
                    elif PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".backPicSize":
                        file.write("\t\t.backPicSize = MON_COORDS_SIZE(" + data.value + "),\n")
        #                continue
                    elif PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".palette":
                        CurrentSpeciesCptl = PkmnDataFile.cell(row = data.row, column = PkmnDataFile.min_column).value
                        CurrentSpeciesCptl = CurrentSpeciesCptl[0] + CurrentSpeciesCptl[1:len(CurrentSpeciesCptl)].lower()
                        file.write("\t\t.palette = gMonPalette_" + CurrentSpeciesCptl  + ",\n")
                    elif PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".shinyPalette":
                        CurrentSpeciesCptl = PkmnDataFile.cell(row = data.row, column = PkmnDataFile.min_column).value
                        CurrentSpeciesCptl = CurrentSpeciesCptl[0] + CurrentSpeciesCptl[1:len(CurrentSpeciesCptl)].lower()
                        file.write("\t\t.shinyPalette = gMonShinyPalette_" + CurrentSpeciesCptl  + ",\n")
                    elif PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == "FOOTPRINT":
                        CurrentSpeciesCptl = PkmnDataFile.cell(row = data.row, column = PkmnDataFile.min_column).value
                        CurrentSpeciesCptl = CurrentSpeciesCptl[0] + CurrentSpeciesCptl[1:len(CurrentSpeciesCptl)].lower()
        #                file.write("\t\tFOOTPRINT(" + CurrentSpeciesCptl  + ")\n")    
                    elif PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".iconSprite":
                        CurrentSpeciesCptl = PkmnDataFile.cell(row = data.row, column = PkmnDataFile.min_column).value
                        CurrentSpeciesCptl = CurrentSpeciesCptl[0] + CurrentSpeciesCptl[1:len(CurrentSpeciesCptl)].lower()
                        file.write("\t\t.iconSprite = gMonIcon_" + CurrentSpeciesCptl  + ",\n")
                    elif PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".levelUpLearnset":
                        CurrentSpeciesCptl = PkmnDataFile.cell(row = data.row, column = PkmnDataFile.min_column).value
                        CurrentSpeciesCptl = CurrentSpeciesCptl[0] + CurrentSpeciesCptl[1:len(CurrentSpeciesCptl)].lower()
        #                file.write("\t\t.levelUpLearnset = s" + CurrentSpeciesCptl  + "LevelUpLearnset,\n")
                    elif PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".teachableLearnset":
                        CurrentSpeciesCptl = PkmnDataFile.cell(row = data.row, column = PkmnDataFile.min_column).value
                        CurrentSpeciesCptl = CurrentSpeciesCptl[0] + CurrentSpeciesCptl[1:len(CurrentSpeciesCptl)].lower()
        #                file.write("\t\t.teachableLearnSet = s" + CurrentSpeciesCptl  + "TeachableLearnset,\n")
        #             elif PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".eggMoveLearnset":
        #                 CurrentSpeciesCptl = PkmnDataFile.cell(row = data.row, column = PkmnDataFile.min_column).value
        #                 CurrentSpeciesCptl = CurrentSpeciesCptl[0] + CurrentSpeciesCptl[1:len(CurrentSpeciesCptl)].lower()
        #                 file.write("\t\t.eggMoveLearnset = s" + CurrentSpeciesCptl  + "TeachableLearnset,\n")
                    elif PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".evolutions" and data.value != None:
                        file.write("\t\t.evolutions = EVOLUTION({EVO_LEVEL, " + data.value  + ", SPECIES_" + PkmnDataFile.cell(data.row + 1, PkmnDataFile.min_column).value + "}),\n")
                    elif PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == "newspecies":
                        continue
                    elif PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".natDexNeeded":
                        continue
                    elif not data.value:
                        continue
                    else:
                        file.write("\t\t" + SpeciesStructAttributes[data.column-1] + " = " + str(data.value) + ",\n")
                
            #close this species
            if species[PkmnDataFile.max_column-2].value == 1:#species tuple is 0 indexed; maxcol is 1 indexed
                file.write("\t},\n\n")
        
        end = time.time()
        print(f"Time: {end - start} seconds")
        
        file.write("#ifdef __INTELLISENSE__\n")
        file.write("};\n")
        file.write("#endif\n")
        
        file.write("//end of program")
