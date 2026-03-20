#Create moveset
import time
import openpyxl as pyxl

from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

start = time.time()

#Globals for making header, opening data, debug prints, etc
Debug = 1
WriteOrAdd = 'w'
GenName = "PkmnEvolved"
CurrentFamily = ""
CurrentSpecies = ""
CurrentSpeciesCptl = ""

PkmnData = load_workbook('pkmndata.xlsx')
if 'tutor-set' in PkmnData.sheetnames:
    PkmnDataFile = PkmnData['tutor-set']
else:
    print("tutor-set sheet not found")

Header = "{\n"

#Print high level information about datafile being accessed
if Debug:
    print(f"First row for species {PkmnDataFile.min_row}")
    print(f"Last row for species {PkmnDataFile.max_row}")
    print(f"First column of species-data ")
    print(f"Last column of species-data  ")
    print(f"First column of tutor-data #{PkmnDataFile.min_column}, Letter:{get_column_letter(PkmnDataFile.min_column)}")
    print(f"Last column of tutor-data  #{PkmnDataFile.max_column}, Letter:{get_column_letter(PkmnDataFile.max_column)}")

with open("all_learnables.json", WriteOrAdd) as file:
    file.write(Header)
    #Start from second row so you do not grab data headers
    #for row in PkmnDataFile.iter_rows(min_row=2, max_row=20, min_col=PkmnDataFile.min_column, max_col=PkmnDataFile.max_column):
    for row in PkmnDataFile.iter_rows(min_row=2, max_row=PkmnDataFile.max_row, min_col=PkmnDataFile.min_column, max_col=PkmnDataFile.max_column):
        for data in row:
            if data.column == 1: #Check if this is a new mon
                CurrentSpecies = data.value
                file.write(f"\t\"{CurrentSpecies}\": [\n")
           
            elif data.value == None or data.value == 1: #if reached newspecies column, do not write 1. check if next species is new species and close with #endif
                file.write(f"\t],\n")
                break #break from the data for loop because you found the last move
            
            else:
                if PkmnDataFile.cell(data.row, data.column + 1).value == None or PkmnDataFile.cell(data.row, data.column + 1).value == 1:
                    file.write(f"\t\"MOVE_{data.value}\"\n")
                else:
                    file.write(f"\t\"MOVE_{data.value}\",\n")
    
    file.write(f"}}")  #close out cause above has error for final mon entry  
    end = time.time()
    print(f"Runtime: {end - start} seconds")
