#Tool for creating new gen.h files for emerald expansion
#Takes information from excel sheet and prepares .h file for custom pokemon
import os
import re
import shutil
import openpyxl as pyxl

from openpyxl.workbook import Workbook
from openpyxl import load_workbook

Debug = 0

FrontSprite = "front.png"
BackSprite = "back.png"
Icon = "icon.png"
InputFolderPath = "Sprites-MASTER/"
FrontSpritePath = "FrontSprite"
FrontSpriteShinyPath = "FrontSpriteShiny"
BackSpritePath = "BackSprite"
BackSpriteShinyPath = "BackSpriteShiny"
IconPath = "Icon"
OutputFolderPath = "graphics"

SpecialCases = ["NIDORAN_M",#Nidoran_m
                "NIDORAN_F",#Nidoran_f
                "MR_MIME",#mr mime
                "MR_JEST",#mr jest
                "MR_FOOL"]#mr fool

#Globals for making header, opening data, debug prints, etc
PkmnData = load_workbook('pkmndata.xlsx')
PkmnDataFile = PkmnData['sanity-data']

print(f"FrontSprite Dir: {os.path.isdir(InputFolderPath + FrontSpritePath)}")
print(f"BackSprite Dir: {os.path.isdir(InputFolderPath + BackSpritePath)}")
print(f"FrontSpriteShiny Dir: {os.path.isdir(InputFolderPath + FrontSpriteShinyPath)}")
print(f"BackSpriteShiny Dir: {os.path.isdir(InputFolderPath + BackSpriteShinyPath)}")
print(f"Icon Dir: {os.path.isdir(InputFolderPath + IconPath)}\n")
os.makedirs(OutputFolderPath, exist_ok = True)

#Step through source file and see if sprite files have a matching sprite
#for species in PkmnDataFile.iter_rows(min_row=2, max_row=2, min_col=1, max_col=1):
if Debug == 0:
    for row in PkmnDataFile.iter_rows(min_row=2, max_row=PkmnDataFile.max_row, min_col=1, max_col=1):
    #for row in PkmnDataFile.iter_rows(min_row=126, max_row=127, min_col=1, max_col=1):
    #for row in PkmnDataFile.iter_rows(min_row=2, max_row=6, min_col=1, max_col=1):
        for file in os.listdir(f"{InputFolderPath + FrontSpritePath}"):
            fileSplit = re.split(r"[_.]",file)
            for word in fileSplit:
                if row[0].value == word:
                    os.makedirs(f"{OutputFolderPath}/{row[0].value.lower()}", exist_ok = True)
                    shutil.copy(f"{InputFolderPath + FrontSpritePath}/{file}", f"{OutputFolderPath}/{row[0].value.lower()}/front.png")
                    break
            
#         for file in os.listdir("./Sprites-MASTER/FrontSpriteShiny"):
#             if data.value in file:
#                 os.makedirs(f"graphics/{data.value.lower()}", exist_ok = True)
#                 shutil.copy(f"FrontSpriteShiny/{file}", f"graphics/{data.value.lower()}/front-shiny.png")
#         for file in os.listdir("./Sprites-MASTER/BackSprite"):
#             if data.value in file:
#                 os.makedirs(f"graphics/{data.value.lower()}", exist_ok = True)
#                 shutil.copy(f"BackSprite/{file}", f"graphics/{data.value.lower()}/back.png")
        for file in os.listdir(f"{InputFolderPath + BackSpriteShinyPath}"):#backsprite copy name is still shiny due to palette generation of make
            fileSplit = re.split(r"[_.]",file)
            for word in fileSplit:
                if row[0].value == word:
                    os.makedirs(f"{OutputFolderPath}/{row[0].value.lower()}", exist_ok = True)
                    shutil.copy(f"{InputFolderPath + BackSpriteShinyPath}/{file}", f"{OutputFolderPath}/{row[0].value.lower()}/back.png")
                
        for file in os.listdir(f"{InputFolderPath + IconPath}"):
            fileSplit = re.split(r"[_.]",file)
            #print(fileSplit)
            for word in fileSplit:
                if row[0].value == word:
                    os.makedirs(f"{OutputFolderPath}/{row[0].value.lower()}", exist_ok = True)
                    shutil.copy(f"{InputFolderPath + IconPath}/{file}", f"{OutputFolderPath}/{row[0].value.lower()}/icon.png")
        
        for case in SpecialCases:
            for file in os.listdir(f"{InputFolderPath + FrontSpritePath}"):
                if row[0].value in file:
                    os.makedirs(f"{OutputFolderPath}/{row[0].value.lower()}", exist_ok = True)
                    shutil.copy(f"{InputFolderPath + FrontSpritePath}/{file}", f"{OutputFolderPath}/{row[0].value.lower()}/front.png")
#         for file in os.listdir("./Sprites-MASTER/FrontSpriteShiny"):
#             if data.value in file:
#                 os.makedirs(f"graphics/{data.value.lower()}", exist_ok = True)
#                 shutil.copy(f"FrontSpriteShiny/{file}", f"graphics/{data.value.lower()}/front-shiny.png")
#         for file in os.listdir("./Sprites-MASTER/BackSprite"):
#             if data.value in file:
#                 os.makedirs(f"graphics/{data.value.lower()}", exist_ok = True)
#                 shutil.copy(f"BackSprite/{file}", f"graphics/{data.value.lower()}/back.png")
            for file in os.listdir(f"{InputFolderPath + BackSpriteShinyPath}"):#backsprite copy name is still shiny due to palette generation of make
                if row[0].value in file:
                    os.makedirs(f"{OutputFolderPath}/{row[0].value.lower()}", exist_ok = True)
                    shutil.copy(f"{InputFolderPath + BackSpriteShinyPath}/{file}", f"{OutputFolderPath}/{row[0].value.lower()}/back.png")
                    
            for file in os.listdir(f"{InputFolderPath + IconPath}"):
                if row[0].value in file:
                    os.makedirs(f"{OutputFolderPath}/{row[0].value.lower()}", exist_ok = True)
                    shutil.copy(f"{InputFolderPath + IconPath}/{file}", f"{OutputFolderPath}/{row[0].value.lower()}/icon.png")