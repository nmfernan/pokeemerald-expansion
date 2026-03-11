#Tool for creating new gen.h files for emerald expansion
#Takes information from excel sheet and prepares .h file for custom pokemon
import openpyxl as pyxl

from openpyxl.workbook import Workbook
from openpyxl import load_workbook

#Globals for making header, opening data, debug prints, etc
Debug = 1
WriteOrAdd = 'w'
GenName = "PkmnEvolved"
PkmnData = load_workbook('pkmndata.xlsx')
PkmnDataFile = PkmnData['trainers']

#Header Description for files, etc
Header ="//gen file for " + GenName + " trainers"
        
#with open(GenName+".h", WriteOrAdd) as file:
with open("test_trainers.h", WriteOrAdd) as file:
    #Print high level information about datafile being accessed
    if Debug:
        print(f"First row for trainers {PkmnDataFile.min_row}")
        print(f"Last row for trainers {PkmnDataFile.max_row}")
        print(f"First column of trainers {PkmnDataFile.min_column}")
        print(f"Last column of trainers  {PkmnDataFile.max_column}")
    
    #Write top level information to species file
    file.write(Header +  "\n")

    #Begin writing species information to .h file
    if Debug == 1:
        #Start from second row so you do not grab data headers
        for row in PkmnDataFile.iter_rows(min_row=2, max_row=10, min_col=PkmnDataFile.min_column, max_col=PkmnDataFile.max_column):
        #for species in PkmnDataFile.iter_rows(min_row=2, max_row=PkmnDataFile.max_row, min_col=PkmnDataFile.min_column, max_col=PkmnDataFile.max_column):
            if row[PkmnDataFile.max_column-1].value != None:#species tuple is 0 indexed; maxcol is 1 indexed
                print("New Trainer Found!: " + row[PkmnDataFile.min_column-1].value)
                file.write("===" + row[PkmnDataFile.min_column-1].value + " ===\n")
            #file.write("\t[SPECIES_" + str(species[PkmnDataFile.min_column-1].value) + "] =\n")
            #file.write("\t{\n")