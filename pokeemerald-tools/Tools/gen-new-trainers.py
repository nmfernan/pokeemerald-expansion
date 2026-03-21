#Tool for creating new gen.h files for emerald expansion
#Takes information from excel sheet and prepares .h file for custom pokemon
import time
import math
from enum import Enum

import openpyxl as pyxl
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

#Globals for making header, opening data, debug prints, etc
class Trainer(Enum):
    ENCOUNTER = 0
    UNSURE = 1
    TYPE = 2
    MUSIC = 3
    GENDER = 4
    SPRITE = 5
    NAME = 6
    ITEM1 = 7
    ITEM2 = 8
    ITEM3 = 9
    ITEM4 = 10
    BATTLE_TYPE = 11
    AI = 12
    NUM_MONS = 13
    MUGSHOT = 14

class Mon(Enum):
    IV = 0
    LVL = 1
    SPECIES = 2
    ITEM = 3
    MOVE1 = 4
    MOVE2 = 5
    MOVE3 = 6
    MOVE4 = 7

Debug = 0
WriteOrAdd = 'w'

FileName = "trainers_frlg.party"

PkmnData = load_workbook('trainer-data.xlsx')
SheetName = "low-level-data-final-3"

if SheetName in PkmnData.sheetnames:
    PkmnDataFile = PkmnData[SheetName]
else:
    print(f"No sheet found")
    print(f"{PkmnData.sheetnames}")

Header = """=== TRAINER_NONE ===
Name: PH
Class: Pkmn Trainer 1
Pic: Hiker Frlg
Gender: Male
Music: Male
Double Battle: No

"""

if Debug:
    print(f"First row for species {PkmnDataFile.min_row}")
    print(f"Last row for species {PkmnDataFile.max_row}")
    print(f"First column of species-data {PkmnDataFile.min_column}")
    print(f"Last column of species-data  {PkmnDataFile.max_column}")
    print(f"First column of tutor-data #{PkmnDataFile.min_column}, Letter:{get_column_letter(PkmnDataFile.min_column)}")
    print(f"Last column of tutor-data  #{PkmnDataFile.max_column}, Letter:{get_column_letter(PkmnDataFile.max_column)}")    

start = time.time()
Check = 0

with open(FileName, WriteOrAdd) as file:
    file.write(Header)
    #for row in PkmnDataFile.iter_rows(min_row=2, max_row=5, min_col=PkmnDataFile.min_column, max_col=1):
    for row in PkmnDataFile.iter_rows(min_row=2, max_row=PkmnDataFile.max_row, min_col=PkmnDataFile.min_column, max_col=1):
        data = str(row[0].value)
        data = data.split(",")
        if Debug: print(data)
        if data[0] == "NEW_TRAINER":
            Check = 1
        
        elif Check == 1:
            file.write(f"=== TRAINER_{data[Trainer.TYPE.value]}_{data[Trainer.ENCOUNTER.value]} ===\n")
            file.write(f"Name: {data[Trainer.NAME.value]}\n")
            if "_" in data[Trainer.TYPE.value]:#I goofed during parsing and have to have this handler now
                data[Trainer.TYPE.value] = data[Trainer.TYPE.value].replace("_"," ")
                
            file.write(f"Class: {data[Trainer.TYPE.value].title()} Frlg\n")
            file.write(f"Pic: {data[Trainer.TYPE.value].title()} Frlg\n")
            file.write(f"Gender: {data[Trainer.GENDER.value].capitalize()}\n")
            file.write(f"Music: {data[Trainer.MUSIC.value].capitalize()}\n")
            if data[Trainer.ITEM1.value] != "":
                file.write(f"Items: ")
                for item in data[Trainer.ITEM1.value:Trainer.BATTLE_TYPE.value]:#have to run to BATTLE_TYPE due to not counting ITEM4 case
                    if "_" in item:#I goofed during parsing and have to have this handler now
                        item = item.replace("_"," ")
                    if item != None:
                        file.write(f"{item.capitalize()} / ")
#                        if Debug: print("item1")
                    else:
#                        if Debug: print("no more items")
                        break
            
            file.write(f"Double Battle: {data[Trainer.BATTLE_TYPE.value].capitalize()}\n")
            file.write(f"AI: {data[Trainer.AI.value]}\n")
            
            try:
                file.write(f"Mugshot: {data[Trainer.MUGSHOT.value]}\n")
            except:
                file.write(f"\n")
                Check = 0                
                continue
            
            Check = 0         
        
        else:
#            if Debug: print("new mon")
            file.write(f"{data[Mon.SPECIES.value].capitalize()}")
            if data[Mon.ITEM.value] == "0" or data[Mon.ITEM.value] == "":
                file.write(f"\n")
            else:
                file.write(f" @ {data[Mon.ITEM.value]}\n")
                
            file.write(f"Level: {data[Mon.LVL.value]}\n")
            iv = int(data[Mon.IV.value]) * 31 / 255
            iv = math.floor(iv)
            file.write(f"IVs: {iv} HP / {iv} Atk / {iv} Def / {iv} SpA / {iv} SpD / {iv} Spe\n")
            for move in data[Mon.MOVE1.value:Mon.MOVE4.value + 1]:#have to run to MOVE4 + 1 due to not counting MOVE4 case
                if "_" in move:#I goofed during parsing and have to have this handler now
                    move = move.replace("_"," ")
                if move != "" and move != "-":
                    file.write(f"- {move.title()}\n")
#                    if Debug: print("move1")
                else:
#                    if Debug: print("no more moves")
                    break
            
            file.write(f"\n")

end = time.time()
print(f"{end-start} seconds")
print("Program completed")
