#Tool for creating new gen.h files for emerald expansion
#Takes information from excel sheet and prepares .h file for custom pokemon
import openpyxl as pyxl

from openpyxl.workbook import Workbook
from openpyxl import load_workbook

#Globals for making header, opening data, debug prints, etc
Debug = 1
WriteOrAdd = 'w'
GenName = "PkmnEvolved"
PkmnData = load_workbook('pkmndata.xlsx')
PkmnDataFile = PkmnData.active
whatFix = "egg"

if whatFix == "types":
    for row in PkmnDataFile.iter_rows(min_row=1, max_row=PkmnDataFile.max_row, min_col=1, max_col=PkmnDataFile.max_column):
        for data in row:
            if PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == "type1" or PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == "type2":
                match data.value:
                    case 0:
                        data.value = "NORMAL"
                    case 1:
                        data.value = "FIGHTING"
                    case 2:
                        data.value = "FLYING"
                    case 3:
                        data.value = "POISON"
                    case 4:
                        data.value = "GROUND"
                    case 5:
                        data.value = "ROCK"
                    case 6:
                        data.value = "BUG"
                    case 7:
                        data.value = "GHOST"
                    case 8:
                        data.value = "STEEL"
                    case 9:
                        data.value = "?????"
                    case 10:
                        data.value = "FIRE"
                    case 11:
                        data.value = "WATER"
                    case 12:
                        data.value = "GRASS"
                    case 13:
                        data.value = "ELECTRIC"
                    case 14:
                        data.value = "PSYCHIC"
                    case 15:
                        data.value = "ICE"
                    case 16:
                        data.value = "DRAGON"
                    case 17:
                        data.value = "DARK"

elif whatFix == "gender":
    for row in PkmnDataFile.iter_rows(min_row=1, max_row=PkmnDataFile.max_row, min_col=1, max_col=PkmnDataFile.max_column):
        for data in row:
            if PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".genderRatio":
                if data.value == "Genderless":
                    data.value = "MON_GENDERLESS"
                elif data.value == "100% Male":
                    data.value = "MON_MALE"
                elif data.value == "100% Female":
                    data.value = "MON_FEMALE"
                elif data.value == "50% Male & 50% Female":
                    data.value = "PERCENT_FEMALE(50)"   
                elif data.value == "75% Female":
                    data.value = "PERCENT_FEMALE(75)"   
                elif data.value == "75% Male":
                    data.value = "PERCENT_FEMALE(25)"                            
                elif data.value == "87% Male":
                    data.value = "PERCENT_FEMALE(12.5)"                            

elif whatFix == "xpRate":
    for row in PkmnDataFile.iter_rows(min_row=1, max_row=PkmnDataFile.max_row, min_col=1, max_col=PkmnDataFile.max_column):
        for data in row:
            if PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".growthRate":
                if data.value == "Fast":
                    data.value = "FAST"
                elif data.value == "Medium Fast":
                    data.value = "MEDIUM_FAST"
                elif data.value == "Medium Slow":
                    data.value = "MEDIUM_SLOW"
                elif data.value == "Slow":
                    data.value = "SLOW"
#doesn't work, idk why, dumb
elif whatFix == "egg":
    for row in PkmnDataFile.iter_rows(min_row=1, max_row=PkmnDataFile.max_row, min_col=1, max_col=PkmnDataFile.max_column):
        for data in row:
            if PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == "egg1" or PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == "egg2":
                print(data.value)
                if data.value != "egg1" or data.value!= "egg2":
                    data.value = "EGG_GROUP_" + data.value.capitalize()
                    print(data.value)
PkmnData.save("pkmndata-fix.xlsx")
PkmnData.close()
print("Program completed")