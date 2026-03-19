#Tool for creating new gen.h files for emerald expansion
#Takes information from excel sheet and prepares .h file for custom pokemon
import openpyxl as pyxl

from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

#Globals for making header, opening data, debug prints, etc
Debug = 1
WriteOrAdd = 'w'

FileName = "encounter-data-fix.xlsx"
GenName = "PkmnEvolved"

PkmnData = load_workbook('encounter-data-evolved.xlsx')
SheetName = "encounter-data-low-level-parse"
#SheetName = "next-data"


if SheetName in PkmnData.sheetnames:
    PkmnDataFile = PkmnData[SheetName]
else:
    print(f"No sheet found")
    print(f"{PkmnData.sheetnames}")

#whatFix = "encounter-data"
whatFix = "encounter-data"

if Debug:
    print(f"First row for species {PkmnDataFile.min_row}")
    print(f"Last row for species {PkmnDataFile.max_row}")
    print(f"First column of species-data {PkmnDataFile.min_column}")
    print(f"Last column of species-data  {PkmnDataFile.max_column}")
    print(f"First column of tutor-data #{PkmnDataFile.min_column}, Letter:{get_column_letter(PkmnDataFile.min_column)}")
    print(f"Last column of tutor-data  #{PkmnDataFile.max_column}, Letter:{get_column_letter(PkmnDataFile.max_column)}")    

if whatFix == "encounter-data":
    for row in PkmnDataFile.iter_rows(min_row=PkmnDataFile.min_row, max_row=PkmnDataFile.max_row, min_col=PkmnDataFile.min_column, max_col=PkmnDataFile.max_column):
    #for row in PkmnDataFile.iter_rows(min_row=PkmnDataFile.min_row, max_row=5, min_col=PkmnDataFile.min_column, max_col=PkmnDataFile.max_column):
        data = str(row[0].value)
        if "@{" in data and "@}" in data:
            row[0].value = data[:data.find("@}")]
            print(data[:data.find("@}")])
            if Debug: print(row[0].value)
            
            PkmnDataFile.insert_rows(row[0].row + 1, amount = 1)
            data = data[data.find("@{"):]
            data = data.replace("@{","")
            PkmnDataFile.cell(row[0].row + 1, 1).value = data

        elif "@{" in data:
            row[0].value = data[:data.find("@{")]
            if Debug: print(data[:data.find("@{")])
            
            PkmnDataFile.insert_rows(row[0].row + 1, amount = 1)
            data = data[data.find("@{"):]
            data = data.replace("@{","")
            PkmnDataFile.cell(row[0].row + 1, 1).value = data

        elif "@}" in data:
            row[0].value = data.replace("@}","")
            if Debug: print(row[0].value)
            PkmnDataFile.insert_rows(row[0].row + 1, amount = 1)
            PkmnDataFile.cell(row[0].row + 1, 1).value = "NEW ROUTE"
            
#            if "%" in data:            
#                PkmnDataFile.cell(row[0].row + 1 , 1).value = data[data.find("%")-2:]

if whatFix == "next-data":
    for row in PkmnDataFile.iter_rows(min_row=PkmnDataFile.min_row, max_row=PkmnDataFile.max_row, min_col=PkmnDataFile.min_column, max_col=PkmnDataFile.max_column):
        data = str(row[0].value)
        if "{" in data:
            row[0].value = data[:data.find("{")]
            if Debug: print(row[0].value)
            PkmnDataFile.insert_rows(row[0].row + 1, amount = 1)
            PkmnDataFile.cell(row[0].row + 1, 1).value = data.replace("{","")
            PkmnDataFile.cell(row[0].row + 1 , 1).value = data[data.find("%")-2:]


PkmnData.save(FileName)
PkmnData.close()
print("Program completed")