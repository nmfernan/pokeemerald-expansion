#Tool for parsing pokedex entries
import openpyxl as pyxl

from openpyxl.workbook import Workbook
from openpyxl import load_workbook

#Globals for making header, opening data, debug prints, etc
Debug = 1
WriteOrAdd = 'w'
GenName = "PkmnEvolved"
PkmnData = load_workbook('pokedex-data.xlsx')
PkmnDataFile = PkmnData['Sheet3']
        
#Print high level information about datafile being accessed
if Debug:
    print(f"First row for species {PkmnDataFile.min_row}")
    print(f"Last row for species {PkmnDataFile.max_row}")
    print(f"First column of species-data {PkmnDataFile.min_column}")
    print(f"Last column of species-data  {PkmnDataFile.max_column}")

#Begin writing species information to .h file
if Debug == 1:    
    #Start from second row so you do not grab data headers
    #for row in PkmnDataFile.iter_rows(min_row=1, max_row=13, min_col=PkmnDataFile.min_column, max_col=1):
    for row in PkmnDataFile.iter_rows(min_row=1, max_row=PkmnDataFile.max_row, min_col=1, max_col=1):
        for data in row:
            #print(data.value)
            if data.value[0] != "+":
                continue
            elif data.value == None:
                break
            else:
                print(f"Found new mon {data.value}")
                PkmnDataFile.cell(data.row, data.column + 1).value = data.value
                for nextrows in range(1,3):
                    if PkmnDataFile.cell(data.row + nextrows, data.column).value == None:
                        break
                    elif PkmnDataFile.cell(data.row + nextrows, data.column).value [0] == "+":
                        break
                    else:
                        PkmnDataFile.cell(data.row, data.column + 1).value = PkmnDataFile.cell(data.row, data.column + 1).value + PkmnDataFile.cell(data.row + nextrows, data.column).value 

PkmnData.save("dex-entry-fix.xlsx")
PkmnData.close()