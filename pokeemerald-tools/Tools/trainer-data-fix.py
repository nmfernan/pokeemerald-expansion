#Tool for creating new gen.h files for emerald expansion
#Takes information from excel sheet and prepares .h file for custom pokemon
import time
import openpyxl as pyxl

from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

#Globals for making header, opening data, debug prints, etc
Debug = 0
WriteOrAdd = 'w'

FileName = "trainer-data-fix.xlsx"
GenName = "PkmnEvolved"

PkmnData = load_workbook('trainer-data.xlsx')
#SheetName = "low-level-data-raw"
SheetName = "low-level-data-working"

if SheetName in PkmnData.sheetnames:
    PkmnDataFile = PkmnData[SheetName]
else:
    print(f"No sheet found")
    print(f"{PkmnData.sheetnames}")

#whatFix = "encounter-data"
whatFix = "encounter-data"

start = time.time()

if Debug:
    print(f"First row for species {PkmnDataFile.min_row}")
    print(f"Last row for species {PkmnDataFile.max_row}")
    print(f"First column of species-data {PkmnDataFile.min_column}")
    print(f"Last column of species-data  {PkmnDataFile.max_column}")
    print(f"First column of tutor-data #{PkmnDataFile.min_column}, Letter:{get_column_letter(PkmnDataFile.min_column)}")
    print(f"Last column of tutor-data  #{PkmnDataFile.max_column}, Letter:{get_column_letter(PkmnDataFile.max_column)}")    

for row in PkmnDataFile.iter_rows(min_row=2, max_row=PkmnDataFile.max_row, min_col=PkmnDataFile.min_column, max_col=1):
#for row in PkmnDataFile.iter_rows(min_row=2, max_row=10, min_col=PkmnDataFile.min_column, max_col=1):
    data = str(row[0].value)
    if "@{" in data:
        row[0].value = data[:data.find("@{")]
        if Debug: print(row[0].value)
        
        PkmnDataFile.insert_rows(row[0].row + 1, amount = 1)
        data = data[data.find("@{"):]
        data = data.replace("@{","")
        PkmnDataFile.cell(row[0].row + 1, 1).value = data

#     elif "@{" in data:
#         row[0].value = data[:data.find("@{")]
#         if Debug: print(data[:data.find("@{")])
#         
#         PkmnDataFile.insert_rows(row[0].row + 1, amount = 1)
#         data = data[data.find("@{"):]
#         data = data.replace("@{","")
#         PkmnDataFile.cell(row[0].row + 1, 1).value = data

    elif "@}" in data:
        row[0].value = data.replace("@}","")
        if Debug: print(row[0].value)
        PkmnDataFile.insert_rows(row[0].row + 1, amount = 1)
        PkmnDataFile.cell(row[0].row + 1, 1).value = "NEW_TRAINER"

PkmnData.save(FileName)
PkmnData.close()
print("Program completed")