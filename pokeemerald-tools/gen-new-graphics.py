#bulk prep text for sprites
import openpyxl as pyxl

from openpyxl.workbook import Workbook
from openpyxl import load_workbook

Debug = 1
OnlyNewSpecies = 0
WriteOrAdd = 'w'
Anim = False
Footprint = False
GenName = "pkmnevolved"
PkmnData = load_workbook('pkmndata.xlsx')
PkmnDataFile = PkmnData['sanity-data']
CurrentSpecies = ""
CurrentSpeciesCptl = ""
CurrentFamily = ""

with open("pokemon.h", WriteOrAdd) as file:
    if Debug == 1:
        file.write("// Normally, INCBIN_COMP acts like INCBIN_U32, but appends `.lz` to the file, compressing it;\n")
        file.write("// If not compressing overworld gfx, make this an alias to INCBIN_32, so gfx will *not* be compressed\n")
        file.write("#if !(OW_GFX_COMPRESS)\n")
        file.write("#define INCBIN_COMP INCBIN_U32\n")
        file.write("#endif\n\n")
        file.write("const u32 gMonFrontPic_CircledQuestionMark[] = INCBIN_U32(\"graphics/pokemon/question_mark/circled/anim_front.4bpp.smol\");\n")
        file.write("const u32 gMonBackPic_CircledQuestionMark[] = INCBIN_U32(\"graphics/pokemon/question_mark/circled/back.4bpp.smol\");\n")
        file.write("const u16 gMonPalette_CircledQuestionMark[] = INCBIN_U16(\"graphics/pokemon/question_mark/circled/normal.gbapal\");\n")
        file.write("const u16 gMonShinyPalette_CircledQuestionMark[] = INCBIN_U16(\"graphics/pokemon/question_mark/circled/shiny.gbapal\");\n")
        file.write("const u8 gMonIcon_QuestionMark[] = INCBIN_U8(\"graphics/pokemon/question_mark/icon.4bpp\");\n")
        file.write("#if P_FOOTPRINTS\n")
        file.write("#if !P_GBA_STYLE_SPECIES_FOOTPRINTS\n")
        file.write("\tconst u8 gMonFootprint_QuestionMark[] = INCBIN_U8(\"graphics/pokemon/question_mark/footprint.1bpp\");\n")
        file.write("#else\n")
        file.write("\tconst u8 gMonFootprint_QuestionMark[] = INCBIN_U8(\"graphics/pokemon/question_mark/footprint_gba.1bpp\");\n")
        file.write("#endif //P_GBA_STYLE_SPECIES_FOOTPRINTS\n")
        file.write("#endif //P_FOOTPRINTS\n")
        file.write("const u32 gObjectEventPic_Substitute[] = INCBIN_COMP(\"graphics/pokemon/question_mark/overworld.4bpp\");\n\n")

    #for row in PkmnDataFile.iter_rows(min_row=2, max_row=13, min_col=1, max_col=PkmnDataFile.max_column):
    
    if OnlyNewSpecies == 0:
        for row in PkmnDataFile.iter_rows(min_row=2, max_row=PkmnDataFile.max_row, min_col=1, max_col=PkmnDataFile.max_column):
            
            CurrentSpeciesCptl = row[0].value.capitalize()
            print(CurrentSpeciesCptl)
            
            if row[PkmnDataFile.max_column - 1].value == 1:
                CurrentFamily = row[0].value
                if CurrentFamily == "NIDORAN_F":
                    CurrentFamily = "NIDORAN"
                elif CurrentFamily == "HITMONLEE":
                    CurrentFamily = "HITMONS"
                elif CurrentFamily == "KANGAKID":
                    CurrentFamily = "KANGASKHAN"
                file.write(f"#if P_FAMILY_{CurrentFamily}\n")
            
            if Anim == True:
                file.write(f"\t const u32 gMonFrontPic_{CurrentSpeciesCptl}[] = INCBIN_U32(\"graphics/pokemon/{CurrentSpeciesCptl.lower()}/anim_front.4bpp.smol\");\n")
            else:
                file.write(f"\t const u32 gMonFrontPic_{CurrentSpeciesCptl}[] = INCBIN_U32(\"graphics/pokemon/{CurrentSpeciesCptl.lower()}/front.4bpp.smol\");\n")
            
            file.write(f"\t const u32 gMonBackPic_{CurrentSpeciesCptl}[] = INCBIN_U32(\"graphics/pokemon/{CurrentSpeciesCptl.lower()}/back.4bpp.smol\");\n")
            file.write(f"\t const u16 gMonPalette_{CurrentSpeciesCptl}[] = INCBIN_U16(\"graphics/pokemon/{CurrentSpeciesCptl.lower()}/front.gbapal\");\n")
            file.write(f"\t const u16 gMonShinyPalette_{CurrentSpeciesCptl}[] = INCBIN_U16(\"graphics/pokemon/{CurrentSpeciesCptl.lower()}/back.gbapal\");\n")
            file.write(f"\t const u8 gMonIcon_{CurrentSpeciesCptl}[] = INCBIN_U8(\"graphics/pokemon/{CurrentSpeciesCptl.lower()}/icon.4bpp.smol\");\n")
            if Footprint == True:
                file.write("\t const u8 gMonFootprint_{CurrentSpeciesCptl.lower()}[] = INCBIN_U8(\"graphics/pokemon/{CurrentSpeciesCptl.lower()}/footprint.1bpp\");\n")
            
            file.write("\n")
                  
            if PkmnDataFile.cell(row[0].row + 1, PkmnDataFile.max_column).value == 1:#if the next mon is a new species
                file.write(f"#endif //P_FAMILY_{CurrentFamily}\n\n")
            elif PkmnDataFile.cell(row[0].row + 1, PkmnDataFile.min_column).value == None:#handle missingno case?
                file.write(f"#endif //P_FAMILY_{CurrentSpecies}\n\n")
    
    else:
        #for row in PkmnDataFile.iter_rows(min_row=2, max_row=PkmnDataFile.max_row, min_col=1, max_col=PkmnDataFile.max_column):
        for row in PkmnDataFile.iter_rows(min_row=2, max_row=6, min_col=1, max_col=PkmnDataFile.max_column):
            
            CurrentSpeciesCptl = row[0].value.capitalize()
            if row[PkmnDataFile.max_column - 2].value == 1:
            
                if Anim == True:
                    file.write("\t const u32 gMonFrontPic_" + CurrentSpeciesCptl + "[] = INCBIN_U32(\"graphics/pokemon/" + CurrentSpeciesCptl.lower() + "/anim_front.4bpp.smol\");\n")
                else:
                    file.write("\t const u32 gMonFrontPic_" + CurrentSpeciesCptl + "[] = INCBIN_U32(\"graphics/pokemon/" + CurrentSpeciesCptl.lower() + "/front.4bpp.smol\");\n")
                
                file.write("\t const u32 gMonBackPic_" + CurrentSpeciesCptl + "[] = INCBIN_U32(\"graphics/pokemon/" + CurrentSpeciesCptl.lower() + "/back.4bpp.smol\");\n")
                file.write("\t const u16 gMonPalette_" + CurrentSpeciesCptl + "[] = INCBIN_U16(\"graphics/pokemon/" + CurrentSpeciesCptl.lower() + "/front.gbapal\");\n")
                file.write("\t const u16 gMonShinyPalette_" + CurrentSpeciesCptl + "[] = INCBIN_U16(\"graphics/pokemon/" + CurrentSpeciesCptl.lower() + "/back.gbapal\");\n")
                file.write("\t const u8 gMonIcon_" + CurrentSpeciesCptl + "[] = INCBIN_U8(\"graphics/pokemon/" + CurrentSpeciesCptl.lower() + "/icon.4bpp\");\n")
                if Footprint == True:
                    file.write("\t const u8 gMonFootprint_" + CurrentSpeciesCptl + "[] = INCBIN_U8(\"graphics/pokemon/" + CurrentSpeciesCptl.lower() + "/footprint.1bpp\");\n")
                file.write("\n")

    file.write("\tconst u32 gMonFrontPic_Egg[] = INCBIN_U32(\"graphics/pokemon/egg/anim_front.4bpp.smol\");\n")
    file.write("\tconst u16 gMonPalette_Egg[] = INCBIN_U16(\"graphics/pokemon/egg/normal.gbapal\");\n")
    file.write("\tconst u8 gMonIcon_Egg[] = INCBIN_U8(\"graphics/pokemon/egg/icon.4bpp\");\n")

#might be needed in future to generate animations, etc
                
#        for data in species:
#            if PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".natDexNeeded" and data.value == 1:
#                CurrentSpeciesCptl = PkmnDataFile.cell(row = data.row, column = PkmnDataFile.min_column).value
#                CurrentSpeciesCptl = CurrentSpeciesCptl[0] + CurrentSpeciesCptl[1:len(CurrentSpeciesCptl)].lower()
#                file.write("SINGLE_ANIMATION(" + CurrentSpeciesCptl + ");\n")
#                file.write("\n")