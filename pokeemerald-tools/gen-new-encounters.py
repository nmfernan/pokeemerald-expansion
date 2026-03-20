#Make Encounters .json
#Takes information from excel sheet and prepares .h file for custom pokemon
import time
import openpyxl as pyxl

from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

start = time.time()

#Globals for execution

Header = """{
  "wild_encounter_groups": [
    {
      "label": "gWildMonHeaders",
      "for_maps": true,
      "fields": [
        {
          "type": "land_mons",
          "encounter_rates": [
            20,
            20,
            10,
            10,
            10,
            10,
            5,
            5,
            4,
            4,
            1,
            1
          ]
        },
        {
          "type": "water_mons",
          "encounter_rates": [
            60,
            30,
            5,
            4,
            1
          ]
        },
        {
          "type": "rock_smash_mons",
          "encounter_rates": [
            60,
            30,
            5,
            4,
            1
          ]
        },
        {
          "type": "fishing_mons",
          "encounter_rates": [
            70,
            30,
            60,
            20,
            20,
            40,
            40,
            15,
            4,
            1
          ],
          "groups": {
            "old_rod": [
              0,
              1
            ],
            "good_rod": [
              2,
              3,
              4
            ],
            "super_rod": [
              5,
              6,
              7,
              8,
              9
            ]
          }
        }
      ],
      "encounters": [
"""
Debug = 1
WriteOrAdd = 'w'
GenName = "Evo"
FileName = "wild_encounters.json"

PkmnData = load_workbook('encounter-data-evolved.xlsx')
SheetName = "final-encounter-data-2"
if SheetName in PkmnData.sheetnames:
    PkmnDataFile = PkmnData[SheetName]
else:
    print(f"No sheet found")
    print(f"{PkmnData.sheetnames}")

#Print high level file information
print(f"First row for species {PkmnDataFile.min_row}")
print(f"Last row for species {PkmnDataFile.max_row}")
print(f"First column of species-data {PkmnDataFile.min_column}")
print(f"Last column of species-data  {PkmnDataFile.max_column}")
print(f"First column of tutor-data #{PkmnDataFile.min_column}, Letter:{get_column_letter(PkmnDataFile.min_column)}")
print(f"Last column of tutor-data  #{PkmnDataFile.max_column}, Letter:{get_column_letter(PkmnDataFile.max_column)}")    

with open(FileName, WriteOrAdd) as file:
    #file.write(Header)
    for row in PkmnDataFile.iter_rows(min_row=1, max_row=PkmnDataFile.max_row, min_col=PkmnDataFile.min_column, max_col=PkmnDataFile.max_column):
    #for row in PkmnDataFile.iter_rows(min_row=1, max_row=915, min_col=PkmnDataFile.min_column, max_col=PkmnDataFile.max_column):
        if row[0].value != None:        
            data = row[0].value.split(",")
        else:
            continue
        #if Debug: print(row[0].value)
        #if Debug: print(data)
        if "NEW_ROUTE" in data:
            file.write(f"\t\t{{\n")
            file.write(f"\t\t\"map\": \"MAP_{data[1]}\",\n")
            file.write(f"\t\t\"base_label\": \"{data[2]}\",\n")
        elif "land_mons" in data:
            file.write(f"\t\t\"{data[0]}\": {{\n")
            file.write(f"\t\t\t\"encounter_rate\": {data[1]},\n")
            file.write(f"\t\t\t\"mons\": [\n")
        elif "water_mons" in data:
            file.write(f"\t\t\"{data[0]}\": {{\n")
            file.write(f"\t\t\t\"encounter_rate\": {data[1]},\n")
            file.write(f"\t\t\t\"mons\": [\n")
        elif "rock_smash_mons" in data:
            file.write(f"\t\t\"{data[0]}\": {{\n")
            file.write(f"\t\t\t\"encounter_rate\": {data[1]},\n")
            file.write(f"\t\t\t\"mons\": [\n")
        elif "fishing_mons" in data:
            file.write(f"\t\t\"{data[0]}\": {{\n")
            file.write(f"\t\t\t\"encounter_rate\": {data[1]},\n")
            file.write(f"\t\t\t\"mons\": [\n")
        else:
            file.write(f"\t\t\t\t{{\n")
            file.write(f"\t\t\t\t\t\"min_level\": {data[0]},\n")
            file.write(f"\t\t\t\t\t\"max_level\": {data[1]},\n")
            file.write(f"\t\t\t\t\t\"species\": \"SPECIES_{data[2]}\"\n")
            closing = PkmnDataFile.cell(row[0].row + 1,row[0].column).value
            if closing == None:
                file.write(f"\t\t\t\t}}\n")
                file.write(f"\t\t\t\t]\n")
                file.write(f"\t\t\t}}\n")
                file.write(f"\t\t}},\n")
                #file.write(f"\t]\n")
                #file.write(f"}},\n")
            elif "mons" in closing:
                file.write(f"\t\t\t\t}}\n")
                file.write(f"\t\t\t]\n")
                file.write(f"\t\t}},\n")
            elif "NEW_ROUTE" in closing:
                file.write(f"\t\t\t\t}}\n")
                file.write(f"\t\t\t\t]\n")
                file.write(f"\t\t\t}}\n")
                file.write(f"\t\t}},\n")
            else:
                file.write(f"\t\t\t\t}},\n")

    for row in PkmnDataFile.iter_rows(min_row=1, max_row=PkmnDataFile.max_row, min_col=PkmnDataFile.min_column, max_col=PkmnDataFile.max_column):
    #for row in PkmnDataFile.iter_rows(min_row=1, max_row=915, min_col=PkmnDataFile.min_column, max_col=PkmnDataFile.max_column):
        if row[0].value != None:        
            data = row[0].value.split(",")
        else:
            continue
        #if Debug: print(row[0].value)
        #if Debug: print(data)
        if "NEW_ROUTE" in data:
            file.write(f"\t\t{{\n")
            file.write(f"\t\t\"map\": \"MAP_{data[1]}\",\n")
            file.write(f"\t\t\"base_label\": \"{data[3]}\",\n") #LeafGreen Encounter Gen
        elif "land_mons" in data:
            file.write(f"\t\t\"{data[0]}\": {{\n")
            file.write(f"\t\t\t\"encounter_rate\": {data[1]},\n")
            file.write(f"\t\t\t\"mons\": [\n")
        elif "water_mons" in data:
            file.write(f"\t\t\"{data[0]}\": {{\n")
            file.write(f"\t\t\t\"encounter_rate\": {data[1]},\n")
            file.write(f"\t\t\t\"mons\": [\n")
        elif "rock_smash_mons" in data:
            file.write(f"\t\t\"{data[0]}\": {{\n")
            file.write(f"\t\t\t\"encounter_rate\": {data[1]},\n")
            file.write(f"\t\t\t\"mons\": [\n")
        elif "fishing_mons" in data:
            file.write(f"\t\t\"{data[0]}\": {{\n")
            file.write(f"\t\t\t\"encounter_rate\": {data[1]},\n")
            file.write(f"\t\t\t\"mons\": [\n")
        else:
            file.write(f"\t\t\t\t{{\n")
            file.write(f"\t\t\t\t\t\"min_level\": {data[0]},\n")
            file.write(f"\t\t\t\t\t\"max_level\": {data[1]},\n")
            file.write(f"\t\t\t\t\t\"species\": \"SPECIES_{data[2]}\"\n")
            closing = PkmnDataFile.cell(row[0].row + 1,row[0].column).value
            if closing == None:
                file.write(f"\t\t\t\t}}\n")
                file.write(f"\t\t\t\t]\n")
                file.write(f"\t\t\t}}\n")
                file.write(f"\t\t}}\n")
                file.write(f"\t]\n")
                file.write(f"}},\n")
            elif "mons" in closing:
                file.write(f"\t\t\t\t}}\n")
                file.write(f"\t\t\t]\n")
                file.write(f"\t\t}},\n")
            elif "NEW_ROUTE" in closing:
                file.write(f"\t\t\t\t}}\n")
                file.write(f"\t\t\t\t]\n")
                file.write(f"\t\t\t}}\n")
                file.write(f"\t\t}},\n")
            else:
                file.write(f"\t\t\t\t}},\n")