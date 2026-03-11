#Create moveset
import openpyxl as pyxl

from openpyxl.workbook import Workbook
from openpyxl import load_workbook

#Globals for making header, opening data, debug prints, etc
Debug = 1
WriteOrAdd = 'w'
GenName = "PkmnEvolved"
PkmnData = load_workbook('pkmndata.xlsx')
#PkmnDataFile = PkmnData.active
if 'learnset' in PkmnData.sheetnames:
    PkmnDataFile = PkmnData['learnset']
else:
    print("learnset sheet not found")

#Header Description for files, etc
Header ="""//learnset for """ + GenName + """
#define LEVEL_UP_MOVE(lvl, moveLearned) {.move = moveLearned, .level = lvl}
#define LEVEL_UP_END {.move = LEVEL_UP_MOVE_END, .level = 0}

static const struct LevelUpMove sNoneLevelUpLearnset[] = {
    LEVEL_UP_MOVE(1, MOVE_POUND),
    LEVEL_UP_END
};
"""

#with open(GenName+".h", WriteOrAdd) as file:
with open("test_learnset.h", WriteOrAdd) as file:
    file.write(Header)
    #Start from second row so you do not grab data headers
    #for species in PkmnDataFile.iter_rows(min_row=2, max_row=13, min_col=PkmnDataFile.min_column, max_col=PkmnDataFile.max_column):
    #for row in PkmnDataFile.iter_rows(min_row=2, max_row=PkmnDataFile.max_row, min_col=PkmnDataFile.min_column, max_col=3):
    for row in PkmnDataFile.iter_rows(min_row=2, max_row=500, min_col=PkmnDataFile.min_column, max_col=3):
    #for species in PkmnDataFile.iter_rows(min_row=2, max_row=PkmnDataFile.max_row, min_col=PkmnDataFile.min_column, max_col=PkmnDataFile.max_column):
        if row[PkmnDataFile.min_column-1].value != None:#Checks if species name is in first column
            fixCase = str(row[0].value) #idk why Excel is being stupid
            fixCase = fixCase[0] + fixCase[1:len(fixCase)].lower()
            if row[2].value == 1:
                print("New Species Found!: " + row[PkmnDataFile.min_column-1].value)
                file.write("#endif\n\n#if P_FAMILY_" + row[PkmnDataFile.min_column-1].value + "\n")
            else:
                file.write("static const struct LevelUpMove s" + fixCase + "LevelUpLearnset[] = {\n")
        elif row[PkmnDataFile.min_column-1].value == None and row[PkmnDataFile.min_column].value == None:
            file.write("\tLEVEL_UP_END\n};\n\n")    
        else:
            file.write("\tLEVEL_UP_MOVE" + row[1].value + ",\n")