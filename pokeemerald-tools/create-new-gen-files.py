#Tool for creating new gen.h files for emerald expansion
#Takes information from excel sheet and prepares .h file for custom pokemon
import openpyxl as pyxl

from openpyxl.workbook import Workbook
from openpyxl import load_workbook

#Globals for making header, opening data, debug prints, etc
Debug = 1
WriteOrAdd = 'w'
GenName = "PkmnEvolved"
PkmnData = load_workbook('pkmndata.xlsx')
PkmnDataFile = PkmnData.active

#Header Description for files, etc
Header ="""//gen file for """ + GenName + """
#ifdef __INTELLISENSE__
const struct SpeciesInfo gSpeciesInfo""" + GenName + """[] =
{
#endif
"""

#Definiton of species information in array format for adding to .h file when iterating through data values
SpeciesStructAttributes = []
for row in PkmnDataFile.iter_rows(min_row=1, max_row=1, min_col=1, max_col=PkmnDataFile.max_column):
    for data in row:
        SpeciesStructAttributes.append(str(data.value))
        print(data.value)
        
#with open(GenName+".h", WriteOrAdd) as file:
with open("test.h", WriteOrAdd) as file:
    #Print high level information about datafile being accessed
    if Debug:
        print(f"First row for species {PkmnDataFile.min_row}")
        print(f"Last row for species {PkmnDataFile.max_row}")
        print(f"First column of species-data {PkmnDataFile.min_column}")
        print(f"Last column of species-data  {PkmnDataFile.max_column}")
    
    #Write top level information to species file
    file.write(Header +  "\n")           

    #Begin writing species information to .h file
    if Debug == 1:
        #Start from second row so you do not grab data headers
        for species in PkmnDataFile.iter_rows(min_row=2, max_row=10, min_col=PkmnDataFile.min_column, max_col=PkmnDataFile.max_column):
        #for species in PkmnDataFile.iter_rows(min_row=2, max_row=PkmnDataFile.max_row, min_col=PkmnDataFile.min_column, max_col=PkmnDataFile.max_column):
            if species[PkmnDataFile.max_column-1].value == 1:#species tuple is 0 indexed; maxcol is 1 indexed
                print("New Species Found!: " + species[PkmnDataFile.min_column-1].value)
                file.write("#if P_FAMILY_" + species[PkmnDataFile.min_column-1].value + "\n")
            file.write("\t[SPECIES_" + str(species[PkmnDataFile.min_column-1].value) + "] =\n")
            file.write("\t{\n")
            
            #step through each element of the species
            for data in species:
                if PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".speciesName":
                    fixCase = data.value
                    fixCase = fixCase[0] + fixCase[1:len(fixCase)].lower()
                    file.write("\t\t" + SpeciesStructAttributes[data.column-1] + " = _(\"" + fixCase + "\"),\n")
                elif PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".types":
                    types = data.value.split(',')
                    type1 = types[0]
                    type2 = types[1]
                    if type1 == type2: #Check for single typing
                        file.write("\t\t.types = MON_TYPES(TYPE_" + type1 + "),\n")
                    else:
                        file.write("\t\t.types = MON_TYPES(TYPE_" + type1 + ", TYPE_"+ type2 + "),\n")    
                elif PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".eggGroups":
                    types = data.value.split(',')
                    type1 = types[0]
                    type2 = types[1]
                    if type1 == type2: #Check for single egg group
                        file.write("\t\t.eggGroups = MON_EGG_GROUPS(EGG_GROUP_" + type1 + "),\n")
                    else:
                        file.write("\t\t.eggGroups = MON_EGG_GROUPS(EGG_GROUP_" + type1 + ", EGG_GROUP_"+ type2 + "),\n")    
                elif PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".abilities":
                    types = data.value.split(',')
                    type1 = "ABILITY_" + types[0]
                    type2 = "ABILITY_" + types[1]
                    type3 = "ABILITY_" + types[2]
                    file.write("\t\t.abilities = { " + type1 + ", " + type2 + " , " + type3 + " },\n")
                elif PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".bodyColor":
                    file.write("\t\t.bodyColor = BODY_COLOR_" + data.value + ",\n")
                elif PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".categoryName":
                    fixCase = data.value
                    fixCase = fixCase[0] + fixCase[1:len(fixCase)].lower()
                    file.write("\t\t" + SpeciesStructAttributes[data.column-1] + " = _(\"" + fixCase + "\"),\n")
                elif PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".description":
                    fixCase = data.value
                    fixCase = fixCase[0] + fixCase[1:len(fixCase)].lower()
                    file.write("\t\t.description = COMPOUD_STRING(\n\t\t\t" + data.value + "),\n")
                elif PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".frontPic":
                    fixCase = PkmnDataFile.cell(row = data.row, column = PkmnDataFile.min_column).value
                    fixCase = fixCase[0] + fixCase[1:len(fixCase)].lower()
                    file.write("\t\t.frontPic = gMonFrontPic_" + fixCase  + ",\n")
                elif PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".frontPicSize":
                    file.write("\t\t.frontPicSize = MON_COORDS_SIZE(" + data.value + "),\n")
                elif PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".backPic":
                    fixCase = PkmnDataFile.cell(row = data.row, column = PkmnDataFile.min_column).value
                    fixCase = fixCase[0] + fixCase[1:len(fixCase)].lower()
                    file.write("\t\t.backPic = gMonBackPic_" + fixCase  + ",\n")
                elif PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".backPic":
                    file.write("\t\t.backPicSize = MON_COORDS_SIZE(" + data.value + "),\n")
                elif PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".palette":
                    fixCase = PkmnDataFile.cell(row = data.row, column = PkmnDataFile.min_column).value
                    fixCase = fixCase[0] + fixCase[1:len(fixCase)].lower()
                    file.write("\t\t.palette = gMonPalette_" + fixCase  + ",\n")
                elif PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".shinyPalette":
                    fixCase = PkmnDataFile.cell(row = data.row, column = PkmnDataFile.min_column).value
                    fixCase = fixCase[0] + fixCase[1:len(fixCase)].lower()
                    file.write("\t\t.shinyPalette = gMonShinyPalette_" + fixCase  + ",\n")
                elif PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".iconSprite":
                    fixCase = PkmnDataFile.cell(row = data.row, column = PkmnDataFile.min_column).value
                    fixCase = fixCase[0] + fixCase[1:len(fixCase)].lower()
                    file.write("\t\t.iconSprite = gMonIconSprite_" + fixCase  + ",\n")
                elif PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".levelUpLearnSet":
                    fixCase = PkmnDataFile.cell(row = data.row, column = PkmnDataFile.min_column).value
                    fixCase = fixCase[0] + fixCase[1:len(fixCase)].lower()
                    file.write("\t\t.iconSprite = s" + fixCase  + "LevelUpLearnset,\n")
                elif PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".teachableLearnset":
                    fixCase = PkmnDataFile.cell(row = data.row, column = PkmnDataFile.min_column).value
                    fixCase = fixCase[0] + fixCase[1:len(fixCase)].lower()
                    file.write("\t\t.iconSprite = s" + fixCase  + "TeachableLearnset,\n")
                elif PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".teachableLearnset":
                    fixCase = PkmnDataFile.cell(row = data.row, column = PkmnDataFile.min_column).value
                    fixCase = fixCase[0] + fixCase[1:len(fixCase)].lower()
                    file.write("\t\t.iconSprite = s" + fixCase  + "TeachableLearnset,\n")
                elif PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == "newspecies":
                    continue
                elif not data.value:
                    continue
                else:
                    file.write("\t\t" + SpeciesStructAttributes[data.column-1] + " = " + str(data.value) + ",\n")
            #close this species
            file.write("\t\t},\n\n")
    elif Debug == 0:
        for species in PkmnDataFile.rows:
            for data in species:
                if data.max_column.value == 1:
                    print("New Species Found")
                print(data.value)
                
    file.write("//end of program")