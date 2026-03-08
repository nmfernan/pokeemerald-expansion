#Tool for creating new gen.h files for emerald expansion
#Takes information from excel sheet and prepares .h file for custom pokemon
import openpyxl as pyxl

from openpyxl.workbook import Workbook
from openpyxl import load_workbook

#Globals for making header, opening data, debug prints, etc
Debug = 1
WriteOrAdd = 'w'
GenName = "PkmnEvolved"
PkmnData = load_workbook('pkmndata.xlsx')
PkmnDataFile = PkmnData.active

for row in PkmnDataFile.iter_rows(min_row=1, max_row=PkmnDataFile.max_row, min_col=1, max_col=PkmnDataFile.max_column):
    for data in row:
        if data.column == 8 or data.column == 9:
            match data.value:
                case 0:
                    data.value = "NORMAL"
                case 1:
                    data.value = "FIGHTING"
                case 2:
                    data.value = "FLYING"
                case 3:
                    data.value = "POISON"
                case 4:
                    data.value = "GROUND"
                case 5:
                    data.value = "ROCK"
                case 6:
                    data.value = "BUG"
                case 7:
                    data.value = "GHOST"
                case 8:
                    data.value = "STEEL"
                case 9:
                    data.value = "?????"
                case 10:
                    data.value = "FIRE"
                case 11:
                    data.value = "WATER"
                case 12:
                    data.value = "GRASS"
                case 13:
                    data.value = "ELECTRIC"
                case 14:
                    data.value = "PSYCHIC"
                case 15:
                    data.value = "ICE"
                case 16:
                    data.value = "DRAGON"
                case 17:
                    data.value = "DARK"
                    
PkmnData.save("pkmndata-fix.xlsx")
PkmnData.close()