#Create moveset
import openpyxl as pyxl

from openpyxl.workbook import Workbook
from openpyxl import load_workbook

GenName = "PkmnEvolved"

#Header Description for files, etc
Header ="""//learnset for """ + GenName + """
#define LEVEL_UP_MOVE(lvl, moveLearned) {.move = moveLearned, .level = lvl}
#define LEVEL_UP_END {.move = LEVEL_UP_MOVE_END, .level = 0}

static const struct LevelUpMove sNoneLevelUpLearnset[] = {
    LEVEL_UP_MOVE(1, MOVE_POUND),
    LEVEL_UP_END
};

"""

#Globals for making header, opening data, debug prints, etc
CurrentFamily = ""
CurrentSpecies = ""
CurrentSpeciesCptl = ""
Debug = 1
WriteOrAdd = 'w'

PkmnData = load_workbook('pkmndata.xlsx')
if 'learnset' in PkmnData.sheetnames:
    PkmnDataFile = PkmnData['learnset']
else:
    print("learnset sheet not found")

#with open(GenName+".h", WriteOrAdd) as file:
with open("gen_EVO.h", WriteOrAdd) as file:
    file.write(Header)
    #Start from second row so you do not grab data headers
    for row in PkmnDataFile.iter_rows(min_row=2, max_row=PkmnDataFile.max_row, min_col=PkmnDataFile.min_column, max_col=3):
    #for row in PkmnDataFile.iter_rows(min_row=2, max_row=500, min_col=PkmnDataFile.min_column, max_col=3):
        if row[PkmnDataFile.min_column-1].value != None:#Checks if species name is in first column
            CurrentSpeciesCptl = row[0].value.capitalize()
            if row[2].value == 1:
                CurrentFamily = row[0].value
                print(f"New Species Found!: {CurrentFamily}")
                if CurrentFamily == "NIDORAN_F":
                    file.write(f"\n#if P_FAMILY_NIDORAN\n") #per emerald expansion declares
                elif CurrentFamily == "KANGAKID":
                    file.write(f"\n#if P_FAMILY_KANGASKHAN\n")    #base form is kanga per emerald expansion declares
                elif CurrentFamily == "HITMONLEE":
                    file.write(f"\n#if P_FAMILY_HITMONS\n") #dumb decalre thing in emerald expansion    
                else:
                    file.write(f"\n#if P_FAMILY_{CurrentFamily}\n")

                file.write(f"static const struct LevelUpMove s{CurrentSpeciesCptl}LevelUpLearnset[] = {{\n")
            else:
                file.write(f"static const struct LevelUpMove s{CurrentSpeciesCptl}LevelUpLearnset[] = {{\n")
            
            file.write(f"\tLEVEL_UP_MOVE{row[1].value},\n")
            
        elif row[PkmnDataFile.min_column-1].value == None and row[PkmnDataFile.min_column].value == None:
            file.write("\tLEVEL_UP_END\n};\n\n")
            if PkmnDataFile.cell(row[0].row + 1, 3).value == 1:
                file.write(f"#endif//{CurrentFamily}\n\n")

        elif row[0].row == PkmnDataFile.max_row:#if first check fails, see if there is no data at all going forward
            file.write(f"\tLEVEL_UP_MOVE{row[1].value},\n")
            file.write("\tLEVEL_UP_END\n};\n")
            file.write("#endif\n\n")
            
        else:
            file.write(f"\tLEVEL_UP_MOVE{row[1].value},\n")