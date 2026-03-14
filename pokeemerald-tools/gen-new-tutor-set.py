#Create moveset
import time
import openpyxl as pyxl

from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

start = time.time()

#Globals for making header, opening data, debug prints, etc
Debug = 1
WriteOrAdd = 'w'
GenName = "PkmnEvolved"
Header = "static const u16 sNoneTeachableLearnset[] = {\n\t MOVE_UNAVAILABLE,\n};\n\n"
CurrentSpecies = ""
PkmnData = load_workbook('pkmndata.xlsx')
if 'tutor-set' in PkmnData.sheetnames:
    PkmnDataFile = PkmnData['tutor-set']
else:
    print("learnset sheet not found")

#Print high level information about datafile being accessed
if Debug:
    print(f"First row for species {PkmnDataFile.min_row}")
    print(f"Last row for species {PkmnDataFile.max_row}")
    print(f"First column of species-data ")
    print(f"Last column of species-data  ")
    print(f"First column of tutor-data #{PkmnDataFile.min_column}, Letter:{get_column_letter(PkmnDataFile.min_column)}")
    print(f"Last column of tutor-data  #{PkmnDataFile.max_column}, Letter:{get_column_letter(PkmnDataFile.max_column)}")

with open("test_teachable.h", WriteOrAdd) as file:
    file.write(Header)
    #Start from second row so you do not grab data headers
    #for row in PkmnDataFile.iter_rows(min_row=2, max_row=20, min_col=PkmnDataFile.min_column, max_col=PkmnDataFile.max_column):
    for row in PkmnDataFile.iter_rows(min_row=2, max_row=PkmnDataFile.max_row, min_col=PkmnDataFile.min_column, max_col=PkmnDataFile.max_column):
        for data in row:
            if data.column == 1: #If a new mon, make check if header necessary
                if row[PkmnDataFile.max_column-1].value == 1:#check for new species flag
                    CurrentSpecies = data.value
                    fixCase = data.value
                    fixCase = fixCase[0] + fixCase[1:len(fixCase)].lower()
                    file.write(f"#if P_FAMILY_{data.value}\nstatic const u16 s{fixCase}TeachableLearnset[] = {{\n")
                else: #if no new species flag, just make new learnset because it belongs to that family evolution line
                    fixCase = data.value
                    fixCase = fixCase[0] + fixCase[1:len(fixCase)].lower()
                    file.write(f"static const u16 s{fixCase}TeachableLearnset[] = {{\n")
#             elif data.column == PkmnDataFile.max_column: #if reached newspecies column, do not write 1. check if next species is new species and close with #endif
#                 if PkmnDataFile.cell(data.row + 1, PkmnDataFile.max_column).value == 1:#if the next mon is a new species
#                     file.write(f"#endif //P_FAMILY_{CurrentSpecies}\n\n")
#                 elif data.value == 1: #account for case when reading row of new species
#                     file.write(f"}};\n")                
#                 else: #otherwise just close it
#                     file.write(f"}};\n")                
            elif data.value == None: #if reached newspecies column, do not write 1. check if next species is new species and close with #endif
                if PkmnDataFile.cell(data.row + 1, PkmnDataFile.max_column).value == 1:#if the next mon is a new species
                    file.write(f"}};\n#endif //P_FAMILY_{CurrentSpecies}\n\n")             
                else: #otherwise just close it
                    file.write(f"}};\n")
                break
            
            elif data.value == 1: #check if you have hit newspecies indicator
                if PkmnDataFile.cell(data.row + 1, PkmnDataFile.max_column).value == 1:#if the next mon is a new species
                    file.write(f"}};\n#endif //P_FAMILY_{CurrentSpecies}\n\n")             
                else: #otherwise just close it
                    file.write(f"}};\n")
            
            else:
                file.write(f"\tMOVE_{data.value},\n")
    
    end = time.time()
    print(f"Runtime: {end - start} seconds")