#Generate specific region and national dex entries
#Requires you to replace the EXISTING Dex with the regional dex result
#Append completely new pokemon to national dex (just makes it easier so other people can work with your game)

import openpyxl as pyxl

from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

Debug = 0
WriteOrAdd = 'w'
GenName = "pkmnevolved"
PkmnData = load_workbook('pkmndata.xlsx')
PkmnDataFile = PkmnData['sanity-data']
Region = "KANTO"
National = "NATIONAL"
#SpeciesIndex = 1573 #based on mega_glimmora being 1572
SpeciesIndex = 0 #based on making new species.h
IncludeConstantsSpecies = 0
DexUpdate = 1

#Print high level information about datafile being accessed
if Debug:
    print(f"First row for species {PkmnDataFile.min_row}")
    print(f"Last row for species {PkmnDataFile.max_row}")
    print(f"First column of species-data ")
    print(f"Last column of species-data  ")
    print(f"First column of tutor-data #{PkmnDataFile.min_column}, Letter:{get_column_letter(PkmnDataFile.min_column)}")
    print(f"Last column of tutor-data  #{PkmnDataFile.max_column}, Letter:{get_column_letter(PkmnDataFile.max_column)}")

if DexUpdate:
    with open("dex-update.h", WriteOrAdd) as file:
        for row in PkmnDataFile.iter_rows(min_row=2, max_row=PkmnDataFile.max_row, min_col=1, max_col=PkmnDataFile.max_column):
        #for row in PkmnDataFile.iter_rows(min_row=2, max_row=20, min_col=1, max_col=PkmnDataFile.max_column):
            if PkmnDataFile.cell(row[0].row, PkmnDataFile.max_column - 1).value == 1:
                file.write(f"\tNATIONAL_DEX_{row[0].value},\n")
    
        for row in PkmnDataFile.iter_rows(min_row=2, max_row=PkmnDataFile.max_row, min_col=1, max_col=PkmnDataFile.max_column):
        #for row in PkmnDataFile.iter_rows(min_row=2, max_row=20, min_col=1, max_col=PkmnDataFile.max_column):
            file.write(f"\tKANTO_DEX_{row[0].value},\n")    

        for row in PkmnDataFile.iter_rows(min_row=2, max_row=PkmnDataFile.max_row, min_col=1, max_col=PkmnDataFile.max_column):
        #for row in PkmnDataFile.iter_rows(min_row=2, max_row=20, min_col=1, max_col=PkmnDataFile.max_column):
            file.write(f"\tKANTO_TO_NATIONAL({row[0].value}),\n")

if Debug:
    with open("species.h", WriteOrAdd) as file:
        file.write("//Species File Update\n")
        file.write("#ifndef GUARD_CONSTANTS_SPECIES_H\n")
        file.write("#define GUARD_CONSTANTS_SPECIES_H\n\n")
        file.write(f"#define SPECIES_NONE \t\t {SpeciesIndex} \n")
        SpeciesIndex += 1
        
        for species in PkmnDataFile.iter_rows(min_row=2, max_row=PkmnDataFile.max_row, min_col=1, max_col=1):
    #             if PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".natDexNeeded" and data.value == 1:
    #                 file.write(f"\t#define SPECIES_ {PkmnDataFile.cell(row = data.row, column = PkmnDataFile.min_column).value} \t\t {SpeciesIndex} \n")
    #                 SpeciesIndex += 1
            file.write(f"#define SPECIES_{species[0].value}\t\t{SpeciesIndex}\n")
            SpeciesIndex += 1
            if species[0].row == PkmnDataFile.max_row:
                file.write(f"\n#define SPECIES_EGG ({species[0].value} + 1)\n")
                
        file.write(f"#define NUM_SPECIES SPECIES_EGG\n")
        file.write(f"#define SPECIES_SHINY_TAG 5000\n")
        file.write(f"#endif  // GUARD_CONSTANTS_SPECIES_H\n")
        
        file.write("//end of program")          

if Debug:
    with open("species_enabled.h", WriteOrAdd) as file: #species_enabled.h config gen   
        file.write("#ifndef GUARD_CONFIG_POKEDEX_H\n")
        file.write("#define GUARD_CONFIG_POKEDEX_H\n\n")
        file.write("#define P_GEN_1_POKEMON                  FALSE // Generation 1 Pokémon (RGBY)\n")
        file.write("#define P_GEN_2_POKEMON                  FALSE // Generation 2 Pokémon (GSC)\n")
        file.write("#define P_GEN_3_POKEMON                  FALSE // Generation 3 Pokémon (RSE, FRLG)\n")
        file.write("#define P_GEN_4_POKEMON                  FALSE // Generation 4 Pokémon (DPPt, HGSS)\n")
        file.write("#define P_GEN_5_POKEMON                  FALSE // Generation 5 Pokémon (BW, B2W2)\n")
        file.write("#define P_GEN_6_POKEMON                  FALSE // Generation 6 Pokémon (XY, ORAS)\n")
        file.write("#define P_GEN_7_POKEMON                  FALSE // Generation 7 Pokémon (SM, USUM, LGPE)\n")
        file.write("#define P_GEN_8_POKEMON                  FALSE // Generation 8 Pokémon (SwSh, BDSP, LA)\n")
        file.write("#define P_GEN_9_POKEMON                  FALSE // Generation 9 Pokémon (SV)\n")
        file.write("#define P_GEN_EVO_POKEMON                  TRUE // Generation EVO Pokémon (??)\n\n")
        
        file.write("// Setting this to TRUE will add the new evolutions to the Regional Dex.\n")
        file.write("#define P_NEW_EVOS_IN_REGIONAL_DEX       TRUE\n\n")
        
        file.write("// Battle gimmick specific Forms.\n")
        file.write("#define P_MEGA_EVOLUTIONS                TRUE\n")
        file.write("#define P_PRIMAL_REVERSIONS              TRUE // Groudon and Kyogre only.\n")
        file.write("#define P_ULTRA_BURST_FORMS              TRUE // Ultra Necrozma only.\n")
        file.write("#define P_GIGANTAMAX_FORMS               TRUE\n")
        file.write("#define P_TERA_FORMS                     TRUE\n\n")
    
        file.write("#define P_GEN_9_MEGA_EVOLUTIONS          P_MEGA_EVOLUTIONS // Mega Evolutions introduced in Z-A and its DLC\n\n")
        
        file.write("// Fusion forms\n")
        file.write("#define P_FUSION_FORMS                   TRUE\n\n")
        
        file.write("// Regional Forms. Includes Regional Form evolutions, like Sirfetch'd.\n")
        file.write("#define P_REGIONAL_FORMS                 TRUE\n")
        file.write("#define P_ALOLAN_FORMS                   P_REGIONAL_FORMS\n")
        file.write("#define P_GALARIAN_FORMS                 P_REGIONAL_FORMS\n")
        file.write("#define P_HISUIAN_FORMS                  P_REGIONAL_FORMS\n")
        file.write("#define P_PALDEAN_FORMS                  P_REGIONAL_FORMS\n\n")
        
        file.write("// Big groups of forms that aren't always desired when choosing families.\n")
        file.write("#define P_PIKACHU_EXTRA_FORMS            TRUE\n")
        file.write("#define P_COSPLAY_PIKACHU_FORMS          P_PIKACHU_EXTRA_FORMS\n")
        file.write("#define P_CAP_PIKACHU_FORMS              P_PIKACHU_EXTRA_FORMS\n\n")

        file.write("// Cross-generation evolutions. Includes pre-evolutions.\n")
        file.write("#define P_CROSS_GENERATION_EVOS          TRUE\n")
        file.write("#define P_GEN_2_CROSS_EVOS               P_CROSS_GENERATION_EVOS\n")
        file.write("#define P_GEN_3_CROSS_EVOS               P_CROSS_GENERATION_EVOS\n")
        file.write("#define P_GEN_4_CROSS_EVOS               P_CROSS_GENERATION_EVOS\n")
        file.write("//#define P_GEN_5_CROSS_EVOS             // Gen 5 didn't introduce any cross-gen evos.\n")
        file.write("#define P_GEN_6_CROSS_EVOS               P_CROSS_GENERATION_EVOS // Just Sylveon.\n")
        file.write("//#define P_GEN_7_CROSS_EVOS             // Alolan evolutions handled by P_ALOLAN_FORMS.\n")
        file.write("#define P_GEN_8_CROSS_EVOS               P_CROSS_GENERATION_EVOS // Regional evolutions handled by P_GALARIAN_FORMS and P_HISUIAN_FORMS.\n")
        file.write("#define P_GEN_9_CROSS_EVOS               P_CROSS_GENERATION_EVOS // Clodsire handled by P_PALDEAN_FORMS.\n\n")

        file.write("// To disable specific families, replace P_GEN_x_POKEMON with FALSE.\n")
        for row in PkmnDataFile.iter_rows(min_row=2, max_row=PkmnDataFile.max_row, min_col=1, max_col=1):
            if PkmnDataFile.cell(row[0].row, PkmnDataFile.max_column).value == 1:
                file.write(f"#define P_FAMILY_{row[0].value}\t\t\tP_GEN_EVO_POKEMON\n")
        file.write(f"\n#endif //GUARD_CONFIG_SPECIES_ENABLED_H")

if IncludeConstantsSpecies == 1:
     with open("new-mons_species.h", WriteOrAdd) as file: #species_enabled.h config gen   
#         for species in PkmnDataFile.iter_rows(min_row=2, max_row=PkmnDataFile.max_row, min_col=PkmnDataFile.min_column, max_col=PkmnDataFile.max_column):
#             for data in species:
#                 if PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".natDexNeeded" and data.value == 1:
#                     file.write(f"#define P_FAMILY_{species[0].value} \t\t\tP_GEN_EVO_POKEMON\n")
#                            
#         file.write("//Species File Update\n")
#         file.write("#ifndef GUARD_CONSTANTS_SPECIES_H\n")
#         file.write("#define GUARD_CONSTANTS_SPECIES_H\n\n")
#         file.write(f"#define SPECIES_NONE \t\t {SpeciesIndex} \n")
#         
        
        CurrentMaxSpecies = 1572
        SpeciesIndex = CurrentMaxSpecies
        SpeciesIndex += 1
        for species in PkmnDataFile.iter_rows(min_row=2, max_row=PkmnDataFile.max_row, min_col=1, max_col=PkmnDataFile.max_column):
            for data in species:
                if PkmnDataFile.cell(row = 1, column = data.column).value == ".natDexNeeded" and data.value == 1:
                    file.write(f"#define SPECIES_{species[0].value:<20}{SpeciesIndex:>25}\n")
                    SpeciesIndex += 1
        
        file.write("//end of program")       

if Debug:
    with open("pokedex.h", WriteOrAdd) as file: #species_enabled.h config gen   
        
        file.write(f"//National Dex Start\n")
        file.write(f"#ifndef GUARD_CONSTANTS_POKEDEX_H\n")
        file.write(f"#define GUARD_CONSTANTS_POKEDEX_H\n")
        file.write(f"// National Pokédex order\n")
        file.write(f"// These constants are NOT disabled by P_GEN_X_POKEMON to keep pokedex_orders.h clean.\n")
        file.write(f"enum NationalDexOrder\n{{\n")
        file.write(f"\tNATIONAL_DEX_NONE,\n")
        
        for species in PkmnDataFile.iter_rows(min_row=2, max_row=PkmnDataFile.max_row, min_col=1, max_col=1):
            for data in species:
                #if PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".natDexNeeded" and data.value == 1:
                file.write("\t" + National + "_DEX_" + PkmnDataFile.cell(row = data.row, column = PkmnDataFile.min_column).value + ",\n")
            
        file.write("};\n\n")
        file.write(f"#define NATIONAL_DEX_COUNT NATIONAL_DEX_{PkmnDataFile.cell(PkmnDataFile.max_row,PkmnDataFile.min_column).value}\n")
        file.write(f"#define POKEMON_SLOTS_NUMBER NATIONAL_DEX_COUNT + 1\n")
        
        
        file.write("// Kanto Pokédex order\n\n")
        file.write("enum KantoDexOrder\n{\n")
        file.write("\tKANTO_DEX_NONE,\n")
        file.write("\t//" + Region + " Dex Start\n")

        for species in PkmnDataFile.iter_rows(min_row=2, max_row=PkmnDataFile.max_row, min_col=1, max_col=1):
            for data in species:
                if PkmnDataFile.cell(row = PkmnDataFile.min_row, column = data.column).value == ".speciesName":
                    file.write("\t" + Region + "_DEX_" + data.value + ",\n")

        file.write("};\n\n")
        file.write(f"#define KANTO_DEX_COUNT (KANTO_DEX_{PkmnDataFile.cell(PkmnDataFile.max_row,PkmnDataFile.min_column).value} + 1)\n")
        file.write(f"#define HOENN_DEX_COUNT 1\n\n")
        file.write(f"#define REGIONAL_DEX_COUNT (IS_FRLG ? KANTO_DEX_COUNT : HOENN_DEX_COUNT)\n\n")
        
        file.write("#define DECAGRAMS_IN_POUND             4536\n")
        file.write("#define CM_PER_INCH                    2.54\n")
        file.write("#define CM_PER_INCH_FACTOR             (CM_PER_INCH * 100)\n")
        file.write("#define INCHES_IN_FOOT                 12\n")
        file.write("#define INCHES_IN_ONE_AND_HALF_FOOT    (INCHES_IN_FOOT * 1.5)\n")
        file.write("#define INCHES_IN_FOOT_FACTOR          (INCHES_IN_FOOT * 10)\n\n")
        
        file.write("#define WEIGHT_HEIGHT_STR_LEN          16\n")
        file.write("#define WEIGHT_HEIGHT_STR_MEM          (WEIGHT_HEIGHT_STR_LEN * sizeof(u8))\n\n")
        
        file.write("#define DEX_HEADER_X                   96\n")
        file.write("#define DEX_Y_TOP                      57\n")
        file.write("#define DEX_Y_BOTTOM                   73\n")
        file.write("#define DEX_MEASUREMENT_X              129\n\n")
        
        file.write("#define DEX_HGSS_HEADER_X_PADDING      59\n")
        file.write("#define DEX_HGSS_Y_TOP_PADDING         7\n") 
        file.write("#define DEX_HGSS_Y_BOTTOM_PADDING      4\n")
        file.write("#define DEX_HGSS_MEASUREMENT_X_PADDING 51\n\n")

        file.write("enum\n{\n\tDEX_MODE_HOENN,\n")
        file.write("\tDEX_MODE_NATIONAL\n};\n\n")
        
        file.write("enum\n{\n\tFLAG_GET_SEEN,\n")
        file.write("\tFLAG_GET_CAUGHT,\n")
        file.write("\tFLAG_SET_SEEN,\n")
        file.write("\tFLAG_SET_CAUGHT\n};\n\n")
        file.write("#endif")
        
        file.write("//end of program")
